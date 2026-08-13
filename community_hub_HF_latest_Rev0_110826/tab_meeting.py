import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from config import supabase, DB_CONNECTED, PLOT_TYPES, TOTAL_PLOTS, TYPE_MAP, load_activities
from utils import get_occupied_count, load_plots


def _load_layout_rows():
    """Fetch all garden_layout rows (block_name, plot_number, plot_type)."""
    try:
        return supabase.table('garden_layout').select('block_name, plot_number, plot_type').execute().data or []
    except Exception:
        return []


def show_meeting(selected_date):
    st.header("📊 Monthly Meeting Dashboard")
    st.caption("Auto-compiled community statistics — ready for monthly meetings")
    if not DB_CONNECTED:
        st.error("Database not connected")
        return

    # ── Period Selector ──
    today = selected_date
    month_start = today.replace(day=1)
    next_m = month_start + timedelta(days=32)
    month_end = next_m.replace(day=1) - timedelta(days=1)
    c1, c2, c3 = st.columns([2, 2, 3])
    with c1:
        start_date = st.date_input("Period Start", value=month_start, key="mtg_start")
    with c2:
        end_date = st.date_input("Period End", value=month_end, key="mtg_end")
    with c3:
        st.write(""); st.write("")
        if st.button("🔄 Refresh Dashboard", use_container_width=True):
            st.rerun()
    st.info(f"Reporting Period: **{start_date.strftime('%d %b %Y')}** → **{end_date.strftime('%d %b %Y')}**")
    st.divider()

    # ═══ SECTION 1 — ROOF TOP GARDEN (MULTI-BLOCK) ═══
    st.subheader("🌱 Roof Top Garden Statistics")
    plots = load_plots()
    occupied = get_occupied_count()
    layout_rows = _load_layout_rows()
    block_names = sorted(set(r['block_name'] for r in layout_rows)) or ['Block 622']
    total_plots_system = len(layout_rows) or TOTAL_PLOTS
    available = total_plots_system - occupied
    occupancy_rate = occupied / total_plots_system if total_plots_system else 0

    gc1, gc2, gc3, gc4, gc5 = st.columns(5)
    gc1.metric("Total Plots", total_plots_system)
    gc2.metric("Occupied", occupied, delta=f"{occupancy_rate:.1%}", delta_color="normal")
    gc3.metric("Available", available)
    gc4.metric("Occupancy Rate", f"{occupancy_rate:.1%}")
    gc5.metric("Blocks", len(block_names))
    st.progress(occupancy_rate, text=f"Garden Utilization: {occupancy_rate:.1%}")

    # 🔥 Per-block breakdown table
    block_rows = []
    for blk in block_names:
        blk_nums = set(r['plot_number'] for r in layout_rows if r['block_name'] == blk)
        blk_plots = [p for p in plots if (p.get('block_name') or 'Block 622') == blk]
        occ_b = len([p for p in blk_plots if p.get('occupied')])
        paid_b = len([p for p in blk_plots if p.get('occupied') and p.get('paid')])
        block_rows.append({
            "Block": blk, "Plots": len(blk_nums), "Occupied": occ_b,
            "Available": len(blk_nums) - occ_b, "Paid": paid_b
        })
    st.markdown("#### Occupancy by Block")
    st.dataframe(pd.DataFrame(block_rows), use_container_width=True, hide_index=True)

        # 🔥 NEW: Occupancy by Block
    st.markdown("#### Occupancy by Block")
    try:
        layout_rows = supabase.table('garden_layout').select('block_name, plot_number').execute().data or []
    except Exception:
        layout_rows = []
    blk_names = sorted(set(r['block_name'] for r in layout_rows)) or ['Block 622']
    blk_rows = []
    for blk in blk_names:
        blk_nums = set(r['plot_number'] for r in layout_rows if r['block_name'] == blk)
        blk_plots = [p for p in plots if (p.get('block_name') or 'Block 622') == blk]
        total_b = len(blk_nums) or len(blk_plots)
        occ_b = len([p for p in blk_plots if p.get('occupied') and (p['plot_number'] in blk_nums if blk_nums else True)])
        paid_b = len([p for p in blk_plots if p.get('occupied') and p.get('paid')])
        blk_rows.append({"Block": blk, "Plots": total_b, "Occupied": occ_b, "Available": total_b - occ_b, "Paid": paid_b})
    st.dataframe(pd.DataFrame(blk_rows), use_container_width=True, hide_index=True)

    st.markdown("#### Occupancy Breakdown by Plot Type (System-wide)")
    type_rows = []
    for tk, ti in PLOT_TYPES.items():
        to = len([p for p in plots if p.get('plot_type') == tk and p.get('occupied')])
        type_rows.append({
            "Plot Type": f"Type {tk} ({ti['area']} m²)",
            "Occupied": to,
            "Available": ti["total"] - to,
            "Rate": round(to / ti["total"] * 100, 1) if ti["total"] else 0
        })
    type_df = pd.DataFrame(type_rows).set_index("Plot Type")
    st.bar_chart(type_df[["Occupied", "Available"]], use_container_width=True, color=["#d62728", "#2ca02c"])
    st.caption("Red = Occupied | Green = Available")
    st.dataframe(type_df.reset_index().rename(columns={"Rate": "Occupancy %"}), use_container_width=True, hide_index=True)

    st.markdown("#### Garden Activity (Requests & Changes)")
    try:
        reqs = supabase.table('plot_requests').select("*").gte('created_at', str(start_date)).lte('created_at', str(end_date)).execute().data
        if reqs:
            req_df = pd.DataFrame(reqs)
            req_df['date'] = pd.to_datetime(req_df['created_at']).dt.date
            daily = req_df.groupby('date').size().reset_index(name='Requests')
            daily = daily.set_index('date')
            st.area_chart(daily, use_container_width=True, color="#ff7f0e")
        else:
            st.info("No garden requests in this period")
    except Exception:
        st.info("Garden request data unavailable")
    st.divider()

    # ═══ SECTION 2 — ACTIVITY PARTICIPATION (S1-S4) ═══
    st.subheader("🥁 Activity Participation")
    acts = load_activities()
    try:
        all_att = supabase.table('attendance').select("*").gte('date', str(start_date)).lte('date', str(end_date)).execute().data
    except Exception:
        all_att = []

    total_records = 0; unique_p = 0
    s1_total = 0; s2_total = 0; s3_total = 0; s4_total = 0
    act_summary = []; parts = None; parts_df = None
    new_c = 0; reg_c = 0; unsigned = 0

    if all_att:
        att_df = pd.DataFrame(all_att)
        att_df['date'] = pd.to_datetime(att_df['date'])
        total_records = len(att_df)
        unique_p = att_df['participant_id'].nunique()
        s1_total = int(att_df['session_1'].sum())
        s2_total = int(att_df['session_2'].sum())
        s3_total = int(att_df['session_3'].sum()) if 'session_3' in att_df.columns else 0
        s4_total = int(att_df['session_4'].sum()) if 'session_4' in att_df.columns else 0
        both_sessions = int(((att_df['session_1'] == True) & (att_df['session_2'] == True)).sum())
        
        ac1, ac2, ac3, ac4, ac5, ac6, ac7 = st.columns(7)
        ac1.metric("Total Records", total_records)
        ac2.metric("Unique Residents", int(unique_p))
        ac3.metric("Session 1", s1_total)
        ac4.metric("Session 2", s2_total)
        ac5.metric("Session 3", s3_total)
        ac6.metric("Session 4", s4_total)
        ac7.metric("Both (S1+S2)", both_sessions)

        st.markdown("#### Participation by Activity")
        act_summary = []
        for act in acts:
            act_name = act['name']
            ad = att_df[att_df['source'] == act_name]
            if not ad.empty:
                act_summary.append({
                    "Activity": act_name, "Records": len(ad),
                    "Unique Residents": int(ad['participant_id'].nunique()),
                    "Session 1": int(ad['session_1'].sum()),
                    "Session 2": int(ad['session_2'].sum()),
                    "Session 3": int(ad['session_3'].sum()) if 'session_3' in ad.columns else 0,
                    "Session 4": int(ad['session_4'].sum()) if 'session_4' in ad.columns else 0,
                    "Both": int(((ad['session_1'] == True) & (ad['session_2'] == True)).sum())
                })
        if act_summary:
            act_sum_df = pd.DataFrame(act_summary)
            st.dataframe(act_sum_df, use_container_width=True, hide_index=True)
            chart_cols = [c for c in ["Session 1", "Session 2", "Session 3", "Session 4"] if c in act_sum_df.columns]
            chart_df = act_sum_df.set_index("Activity")[chart_cols]
            st.bar_chart(chart_df, use_container_width=True)

        st.markdown("#### Daily Attendance Trend")
        agg_dict = {'participant_id': 'nunique', 'session_1': 'sum', 'session_2': 'sum'}
        if 'session_3' in att_df.columns: agg_dict['session_3'] = 'sum'
        if 'session_4' in att_df.columns: agg_dict['session_4'] = 'sum'
        daily = att_df.groupby(att_df['date'].dt.date).agg(agg_dict).reset_index()
        daily.columns = ['Date', 'Unique Residents'] + [f"Session {i}" for i in range(1, len(daily.columns) - 1)]
        daily = daily.set_index('Date')
        st.line_chart(daily, use_container_width=True)

        st.markdown("#### Resident Composition")
        try:
            parts = supabase.table('participants').select("*").execute().data
            if parts:
                parts_df = pd.DataFrame(parts)
                new_c = int(parts_df['is_new'].sum())
                reg_c = len(parts_df) - new_c
                unsigned = int((~parts_df['indemnity'].fillna(False)).sum())
                rc1, rc2, rc3 = st.columns(3)
                rc1.metric("🆕 New", new_c)
                rc2.metric("⭐ Regular", reg_c)
                rc3.metric("🔴 Unsigned Indemnity", unsigned)
        except Exception:
            pass
    else:
        st.info("No attendance records for the selected period")
    st.divider()

    # ═══ SECTION 3 — ACTIVE ACTIVITIES (Dynamic 1-4) ═══
    st.subheader("📅 Active & Upcoming Activities")
    if acts:
        act_cols = st.columns(min(len(acts), 4))
        for i, act in enumerate(acts):
            with act_cols[i % 4]:
                status = "🟢 Active" if act.get('active') else "⚪ Inactive"
                sessions_html = ""
                for s_idx in range(1, 5):
                    lbl = act.get(f'session_{s_idx}_label', '')
                    if lbl and lbl.strip():
                        sessions_html += f"{lbl}<br/>"
                if not sessions_html:
                    sessions_html = "Session 1"
                card_html = f'<div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 15px; border-radius: 10px; margin-bottom: 10px;"><div style="font-size: 18px; font-weight: bold;">{act["name"]}</div><div style="font-size: 12px; margin-top: 5px; opacity: 0.9;">{sessions_html}</div><div style="font-size: 11px; margin-top: 8px; background: rgba(255,255,255,0.2); padding: 4px 8px; border-radius: 4px; display: inline-block;">{status}</div></div>'
                st.markdown(card_html, unsafe_allow_html=True)
    else:
        st.info("No activities configured")
    st.divider()

    # ═══ SECTION 4 — EXCEL EXPORT ═══
    st.subheader("📊 Export Monthly Meeting Report")
    if st.button("📊 Generate Excel Report", type="primary", use_container_width=True):
        with st.spinner("Generating Excel report..."):
            excel_buffer = generate_excel_report(start_date, end_date, occupied, available, occupancy_rate, type_rows, all_att, total_records, unique_p, s1_total, s2_total, s3_total, s4_total, act_summary, parts, parts_df, new_c, reg_c, unsigned)
            st.success("Excel report generated!")
            st.download_button(
                "📥 Download Excel Report", excel_buffer,
                f"monthly_meeting_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    if all_att:
        csv = att_df.to_csv(index=False).encode('utf-8')
        st.download_button("📊 Download Raw Data (CSV)", csv, f"activity_data_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv", "text/csv")


def _hex_to_arbg(hex_color):
    hex_clean = hex_color.replace('#', '').replace('0x', '')
    return ("00" + hex_clean.upper()) if len(hex_clean) == 6 else hex_clean.upper()


def generate_excel_report(start_date, end_date, occupied, available, occupancy_rate, type_rows, all_att, total_records, unique_p, s1_total, s2_total, s3_total, s4_total, act_summary, parts, parts_df, new_c, reg_c, unsigned):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    thin_border = Border(left=Side(style='thin', color='00CCCCCC'), right=Side(style='thin', color='00CCCCCC'),
                         top=Side(style='thin', color='00CCCCCC'), bottom=Side(style='thin', color='00CCCCCC'))

    # ── Sheet 1: Summary Dashboard ──
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.merge_cells('A1:H1')
    ws1['A1'] = "Woodlands Zone 6 Community Hub - Monthly Meeting Report"
    ws1['A1'].font = Font(size=18, bold=True, color="00FFFFFF")
    ws1['A1'].fill = PatternFill(start_color="00667EEA", end_color="00667EEA", fill_type="solid")
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 30
    
    ws1.merge_cells('A2:H2')
    ws1['A2'] = f"Period: {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}"
    ws1['A2'].font = Font(size=12, italic=True)
    ws1['A2'].alignment = Alignment(horizontal='center')
    
    ws1.merge_cells('A3:H3')
    ws1['A3'] = f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}"
    ws1['A3'].font = Font(size=10, color="00666666")
    ws1['A3'].alignment = Alignment(horizontal='center')

    row = 5
    ws1.merge_cells(f'A{row}:H{row}')
    ws1[f'A{row}'] = "ROOF TOP GARDEN"
    ws1[f'A{row}'].font = Font(size=14, bold=True, color="00FFFFFF")
    ws1[f'A{row}'].fill = PatternFill(start_color="002CA02C", end_color="002CA02C", fill_type="solid")
    row += 1
    
    layout_rows = _load_layout_rows()
    total_plots_system = len(layout_rows) or TOTAL_PLOTS
    garden_metrics = [["Total Plots", total_plots_system], ["Occupied", occupied], ["Available", available], ["Occupancy Rate", f"{occupancy_rate:.1%}"]]
    for label, value in garden_metrics:
        ws1[f'A{row}'] = label; ws1[f'B{row}'] = value
        ws1[f'A{row}'].font = Font(bold=True)
        ws1[f'B{row}'].font = Font(size=12, color="00667EEA", bold=True)
        row += 1

    row += 1
    ws1.merge_cells(f'A{row}:H{row}')
    ws1[f'A{row}'] = "Plot Types Breakdown"
    ws1[f'A{row}'].font = Font(size=12, bold=True)
    row += 1
    headers = ["Plot Type", "Occupied", "Available", "Total", "Rate %"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True, color="00FFFFFF")
        cell.fill = PatternFill(start_color="00333333", end_color="00333333", fill_type="solid")
        cell.border = thin_border
    row += 1
    for tr in type_rows:
        parts_type = tr["Plot Type"].split(" ")
        ptype = parts_type[1] if len(parts_type) > 1 else "B"
        argb_color = _hex_to_arbg(PLOT_TYPES.get(ptype, {}).get('colour', '#667EEA'))
        ws1.cell(row=row, column=1, value=tr["Plot Type"]).fill = PatternFill(start_color=argb_color, end_color=argb_color, fill_type="solid")
        ws1.cell(row=row, column=1).font = Font(color="00FFFFFF", bold=True)
        ws1.cell(row=row, column=2, value=tr["Occupied"])
        ws1.cell(row=row, column=3, value=tr["Available"])
        ws1.cell(row=row, column=4, value=tr["Occupied"] + tr["Available"])
        ws1.cell(row=row, column=5, value=f"{tr['Rate']:.1f}%")
        for c in range(1, 6):
            ws1.cell(row=row, column=c).border = thin_border
            ws1.cell(row=row, column=c).alignment = Alignment(horizontal='center')
        row += 1

    row += 1
    ws1.merge_cells(f'A{row}:H{row}')
    ws1[f'A{row}'] = "ACTIVITY PARTICIPATION"
    ws1[f'A{row}'].font = Font(size=14, bold=True, color="00FFFFFF")
    ws1[f'A{row}'].fill = PatternFill(start_color="00764BA2", end_color="00764BA2", fill_type="solid")
    row += 1
    if all_att:
        activity_metrics = [["Total Records", total_records], ["Unique Residents", int(unique_p)], ["Session 1", s1_total], ["Session 2", s2_total], ["Session 3", s3_total], ["Session 4", s4_total]]
        for label, value in activity_metrics:
            ws1[f'A{row}'] = label; ws1[f'B{row}'] = value
            ws1[f'A{row}'].font = Font(bold=True)
            ws1[f'B{row}'].font = Font(size=12, color="00667EEA", bold=True)
            row += 1
        if act_summary:
            row += 1
            ws1.merge_cells(f'A{row}:H{row}')
            ws1[f'A{row}'] = "By Activity"
            ws1[f'A{row}'].font = Font(size=12, bold=True)
            row += 1
            act_headers = ["Activity", "Records", "Unique Residents", "Session 1", "Session 2", "Session 3", "Session 4", "Both"]
            for col, h in enumerate(act_headers, 1):
                cell = ws1.cell(row=row, column=col, value=h)
                cell.font = Font(bold=True, color="00FFFFFF")
                cell.fill = PatternFill(start_color="00333333", end_color="00333333", fill_type="solid")
                cell.border = thin_border
            row += 1
            for a in act_summary:
                ws1.cell(row=row, column=1, value=a['Activity'])
                ws1.cell(row=row, column=2, value=a['Records'])
                ws1.cell(row=row, column=3, value=a['Unique Residents'])
                ws1.cell(row=row, column=4, value=a['Session 1'])
                ws1.cell(row=row, column=5, value=a['Session 2'])
                ws1.cell(row=row, column=6, value=a.get('Session 3', 0))
                ws1.cell(row=row, column=7, value=a.get('Session 4', 0))
                ws1.cell(row=row, column=8, value=a['Both'])
                for c in range(1, 9):
                    ws1.cell(row=row, column=c).border = thin_border
                    ws1.cell(row=row, column=c).alignment = Alignment(horizontal='center')
                row += 1
    else:
        ws1[f'A{row}'] = "No attendance records for this period."
        ws1[f'A{row}'].font = Font(italic=True, color="00666666")

    row += 2
    ws1.merge_cells(f'A{row}:H{row}')
    ws1[f'A{row}'] = "RESIDENT SNAPSHOT"
    ws1[f'A{row}'].font = Font(size=14, bold=True, color="00FFFFFF")
    ws1[f'A{row}'].fill = PatternFill(start_color="00FF7F0E", end_color="00FF7F0E", fill_type="solid")
    row += 1
    try:
        if parts is not None and parts_df is not None:
            resident_metrics = [["Total Registered", len(parts_df)], ["New", new_c], ["Regular", reg_c], ["Unsigned Indemnity", unsigned]]
            for label, value in resident_metrics:
                ws1[f'A{row}'] = label; ws1[f'B{row}'] = value
                ws1[f'A{row}'].font = Font(bold=True)
                ws1[f'B{row}'].font = Font(size=12, color="00667EEA", bold=True)
                row += 1
    except: pass

    for col_idx in range(1, ws1.max_column + 1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = 18

        # ── Sheet 2: Garden Plot Entitlements (PER-BLOCK) ──
    ws2 = wb.create_sheet("Garden Entitlements")
    ws2.merge_cells('A1:I1')
    ws2['A1'] = "Garden Plot Entitlements (All Blocks)"
    ws2['A1'].font = Font(size=16, bold=True, color="00FFFFFF")
    ws2['A1'].fill = PatternFill(start_color="002CA02C", end_color="002CA02C", fill_type="solid")
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 25
    headers2 = ["Block", "Plot Number", "Plot Type", "Area (sqm)", "Status", "Owner ID", "Owner Name", "Contact", "Paid"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color="00FFFFFF")
        cell.fill = PatternFill(start_color="00333333", end_color="00333333", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    try:
        all_participants = supabase.table('participants').select("*").execute().data
        part_dict = {str(p['id']).lower().strip(): p for p in all_participants}
    except Exception:
        part_dict = {}
    all_plots = load_plots()
    try:
        layout_rows = supabase.table('garden_layout').select('block_name, plot_number, plot_type').execute().data or []
    except Exception:
        layout_rows = []
    blocks = sorted(set(r['block_name'] for r in layout_rows)) or ['Block 622']
    row = 3
    for blk in blocks:
        blk_layout = sorted([r for r in layout_rows if r['block_name'] == blk], key=lambda x: x['plot_number'])
        blk_plots = {p['plot_number']: p for p in all_plots if (p.get('block_name') or 'Block 622') == blk}
        ws2.merge_cells(f'A{row}:I{row}')
        ws2[f'A{row}'] = blk
        ws2[f'A{row}'].font = Font(size=13, bold=True, color="00FFFFFF")
        ws2[f'A{row}'].fill = PatternFill(start_color="00667EEA", end_color="00667EEA", fill_type="solid")
        row += 1
        plot_nums = [r['plot_number'] for r in blk_layout] or list(range(1, 77))
        for pn in plot_nums:
            lr = next((r for r in blk_layout if r['plot_number'] == pn), None)
            plot = blk_plots.get(pn)
            ptype = (plot.get('plot_type') if plot and plot.get('plot_type') else None) \
                or (lr.get('plot_type') if lr else None) or TYPE_MAP.get(pn, 'B')
            area = PLOT_TYPES.get(ptype, PLOT_TYPES['B'])['area']
            if plot and plot.get('occupied'):
                status_text = "Occupied"
                owner_id = plot.get('user_id', '')
                resident = part_dict.get(str(owner_id).lower().strip())
                owner_name = resident['name'] if resident else plot.get('user_name', '')
                contact = resident.get('contact', '') if resident else ''
                paid_text = "Yes" if plot.get('paid') else "No"
                status_fill = PatternFill(start_color="00FFCCCC", end_color="00FFCCCC", fill_type="solid")
                status_font = Font(color="00CC0000", bold=True)
            else:
                status_text = "Available"
                owner_id = ""; owner_name = ""; contact = ""; paid_text = ""
                status_fill = PatternFill(start_color="00CCFFCC", end_color="00CCFFCC", fill_type="solid")
                status_font = Font(color="00006600", bold=True)
            ws2.cell(row=row, column=1, value=blk)
            ws2.cell(row=row, column=2, value=pn)
            ws2.cell(row=row, column=3, value=ptype)
            ws2.cell(row=row, column=4, value=area)
            ws2.cell(row=row, column=5, value=status_text)
            ws2.cell(row=row, column=5).fill = status_fill
            ws2.cell(row=row, column=5).font = status_font
            ws2.cell(row=row, column=6, value=owner_id)
            ws2.cell(row=row, column=7, value=owner_name)
            ws2.cell(row=row, column=8, value=contact)
            ws2.cell(row=row, column=9, value=paid_text)
            for c in range(1, 10):
                ws2.cell(row=row, column=c).border = thin_border
                ws2.cell(row=row, column=c).alignment = Alignment(horizontal='center')
            row += 1
        row += 1
    for col_idx in range(1, ws2.max_column + 1):
        max_length = 0
        column = get_column_letter(col_idx)
        for cell in ws2[column]:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws2.column_dimensions[column].width = min(max_length + 2, 50)

    # ── Sheet 3: Raw Attendance Data ──
    if all_att:
        ws3 = wb.create_sheet("Attendance Data")
        att_df_export = pd.DataFrame(all_att)
        for r_idx, row_data in enumerate(dataframe_to_rows(att_df_export, index=False, header=True), 1):
            for c_idx, value in enumerate(row_data, 1):
                ws3.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    ws3.cell(row=r_idx, column=c_idx).font = Font(bold=True, color="00FFFFFF")
                    ws3.cell(row=r_idx, column=c_idx).fill = PatternFill(start_color="00333333", end_color="00333333", fill_type="solid")
                else:
                    ws3.cell(row=r_idx, column=c_idx).border = thin_border
        for col_idx in range(1, ws3.max_column + 1):
            ws3.column_dimensions[get_column_letter(col_idx)].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer