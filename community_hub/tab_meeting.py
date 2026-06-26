import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from config import supabase, DB_CONNECTED, PLOT_TYPES, TOTAL_PLOTS, TYPE_MAP, load_activities
from utils import get_occupied_count, load_plots

def show_meeting(selected_date):
    st.header("📊 Monthly Meeting Dashboard")
    st.caption("Auto-compiled community statistics — ready for monthly meetings")

    if not DB_CONNECTED:
        st.error("Database not connected")
        return

    # ── Period Selector ──
    today = selected_date
    month_start = today.replace(day=1)
    next_m = month_start + timedelta(days=32)
    month_end = next_m.replace(day=1) - timedelta(days=1)

    c1, c2, c3 = st.columns([2, 2, 3])
    with c1:
        start_date = st.date_input("Period Start", value=month_start, key="mtg_start")
    with c2:
        end_date = st.date_input("Period End", value=month_end, key="mtg_end")
    with c3:
        st.write(""); st.write("")
        if st.button("🔄 Refresh Dashboard", use_container_width=True):
            st.rerun()

    st.info(f"Reporting Period: **{start_date.strftime('%d %b %Y')}** → **{end_date.strftime('%d %b %Y')}**")
    st.divider()

    # ═══════════════════════════════════════════
    #  SECTION 1 — ROOF TOP GARDEN
    # ═══════════════════════════════════════════
    st.subheader("🌱 Roof Top Garden Statistics")

    plots = load_plots()
    occupied = get_occupied_count()
    available = TOTAL_PLOTS - occupied
    occupancy_rate = occupied / TOTAL_PLOTS if TOTAL_PLOTS else 0

    # Big metric cards
    gc1, gc2, gc3, gc4, gc5 = st.columns(5)
    gc1.metric("Total Plots", TOTAL_PLOTS)
    gc2.metric("Occupied", occupied, delta=f"{occupancy_rate:.1%}", delta_color="normal")
    gc3.metric("Available", available)
    gc4.metric("Occupancy Rate", f"{occupancy_rate:.1%}")
    gc5.metric("Utilization", f"{occupied}/{TOTAL_PLOTS}")

    st.progress(occupancy_rate, text=f"Garden Utilization: {occupancy_rate:.1%}")

    # By Plot Type — stacked bar chart
    st.markdown("#### Occupancy Breakdown by Plot Type")
    type_rows = []
    for tk, ti in PLOT_TYPES.items():
        to = len([p for p in plots if p.get('plot_type') == tk and p.get('occupied')])
        type_rows.append({
            "Plot Type": f"Type {tk} ({ti['area']} m²)",
            "Occupied": to,
            "Available": ti["total"] - to,
            "Rate": round(to / ti["total"] * 100, 1) if ti["total"] else 0
        })
    type_df = pd.DataFrame(type_rows).set_index("Plot Type")
    st.bar_chart(type_df[["Occupied", "Available"]], use_container_width=True, color=["#d62728", "#2ca02c"])
    st.caption("Red = Occupied | Green = Available")

    # Type summary table
    st.dataframe(
        type_df.reset_index().rename(columns={"Rate": "Occupancy %}"}),
        use_container_width=True, hide_index=True
    )

    # Garden requests trend
    st.markdown("#### Garden Activity (Requests & Changes)")
    try:
        reqs = supabase.table('plot_requests').select("*").gte('created_at', str(start_date)).lte('created_at', str(end_date)).execute().data
        if reqs:
            req_df = pd.DataFrame(reqs)
            req_df['date'] = pd.to_datetime(req_df['created_at']).dt.date
            daily = req_df.groupby('date').size().reset_index(name='Requests')
            daily = daily.set_index('date')
            st.area_chart(daily, use_container_width=True, color="#ff7f0e")
        else:
            st.info("No garden requests in this period")
    except Exception:
        st.info("Garden request data unavailable")

    st.divider()

    # ═══════════════════════════════════════════
    #  SECTION 2 — ACTIVITY PARTICIPATION
    # ═══════════════════════════════════════════
    st.subheader("🥁 Activity Participation")

    acts = load_activities()
    act_names = [a['name'] for a in acts]

    # Load attendance for period
    try:
        all_att = supabase.table('attendance').select("*").gte('date', str(start_date)).lte('date', str(end_date)).execute().data
    except Exception:
        all_att = []

    if all_att:
        att_df = pd.DataFrame(all_att)
        att_df['date'] = pd.to_datetime(att_df['date'])

        total_records = len(att_df)
        unique_p = att_df['participant_id'].nunique()
        s1_total = int(att_df['session_1'].sum())
        s2_total = int(att_df['session_2'].sum())
        both_sessions = int(((att_df['session_1'] == True) & (att_df['session_2'] == True)).sum())

        ac1, ac2, ac3, ac4, ac5 = st.columns(5)
        ac1.metric("Total Records", total_records)
        ac2.metric("Unique Residents", int(unique_p))
        ac3.metric("Session 1", s1_total)
        ac4.metric("Session 2", s2_total)
        ac5.metric("Both Sessions", both_sessions)

        # By Activity table + chart
        st.markdown("#### Participation by Activity")
        act_summary = []
        for act in acts:
            act_name = act['name']
            ad = att_df[att_df['source'] == act_name]
            if not ad.empty:
                act_summary.append({
                    "Activity": act_name,
                    "Records": len(ad),
                    "Unique Residents": int(ad['participant_id'].nunique()),
                    "Session 1": int(ad['session_1'].sum()),
                    "Session 2": int(ad['session_2'].sum()),
                    "Both": int(((ad['session_1'] == True) & (ad['session_2'] == True)).sum())
                })

        if act_summary:
            act_sum_df = pd.DataFrame(act_summary)
            st.dataframe(act_sum_df, use_container_width=True, hide_index=True)

            chart_df = act_sum_df.set_index("Activity")[["Session 1", "Session 2"]]
            st.bar_chart(chart_df, use_container_width=True)

        # Daily trend
        st.markdown("#### Daily Attendance Trend")
        daily = att_df.groupby(att_df['date'].dt.date).agg({
            'participant_id': 'nunique',
            'session_1': 'sum',
            'session_2': 'sum'
        }).reset_index()
        daily.columns = ['Date', 'Unique Residents', 'Session 1', 'Session 2']
        daily = daily.set_index('Date')
        st.line_chart(daily, use_container_width=True)

        # New vs Regular
        st.markdown("#### Resident Composition")
        try:
            parts = supabase.table('participants').select("*").execute().data
            if parts:
                parts_df = pd.DataFrame(parts)
                new_c = int(parts_df['is_new'].sum())
                reg_c = len(parts_df) - new_c
                unsigned = int((~parts_df['indemnity'].fillna(False)).sum())

                rc1, rc2, rc3 = st.columns(3)
                rc1.metric("🆕 New", new_c)
                rc2.metric("⭐ Regular", reg_c)
                rc3.metric("🔴 Unsigned Indemnity", unsigned)
        except Exception:
            pass
    else:
        st.info("No attendance records for the selected period")

    st.divider()

    # ═══════════════════════════════════════════
    #  SECTION 3 — FUTURE / ACTIVE ACTIVITIES
    # ═══════════════════════════════════════════
    st.subheader("📅 Active & Upcoming Activities")
    if acts:
        act_cols = st.columns(min(len(acts), 4))
        for i, act in enumerate(acts):
            with act_cols[i % 4]:
                status = "🟢 Active" if act.get('active') else "⚪ Inactive"
                s1 = act.get('session_1_label', 'Session 1') or 'Session 1'
                s2 = act.get('session_2_label', '') or ''
                if s2.strip():
                    sessions_html = f"{s1}<br/>{s2}"
                else:
                    sessions_html = s1
                card_html = f'<div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 15px; border-radius: 10px; margin-bottom: 10px;"><div style="font-size: 18px; font-weight: bold;">{act["name"]}</div><div style="font-size: 12px; margin-top: 5px; opacity: 0.9;">{sessions_html}</div><div style="font-size: 11px; margin-top: 8px; background: rgba(255,255,255,0.2); padding: 4px 8px; border-radius: 4px; display: inline-block;">{status}</div></div>'
                st.markdown(card_html, unsafe_allow_html=True)
    else:
        st.info("No activities configured")

    st.divider()

    # ═══════════════════════════════════════════
    #  SECTION 4 — EXCEL EXPORT (replaces HTML report)
    # ═══════════════════════════════════════════
    st.subheader("📊 Export Monthly Meeting Report")

    if st.button("📊 Generate Excel Report", type="primary", use_container_width=True):
        with st.spinner("Generating Excel report..."):
            excel_buffer = generate_excel_report(start_date, end_date, occupied, available, occupancy_rate, type_rows, all_att, total_records, unique_p, s1_total, s2_total, act_summary, parts, parts_df, new_c, reg_c, unsigned)
            st.success("Excel report generated!")
            st.download_button(
                "📥 Download Excel Report",
                excel_buffer,
                f"monthly_meeting_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    if all_att:
        csv = att_df.to_csv(index=False).encode('utf-8')
        st.download_button(
            "📊 Download Raw Data (CSV)",
            csv,
            f"activity_data_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv",
            "text/csv"
        )


def _hex_to_arbg(hex_color):
    """Convert #RRGGBB to aRGB format (00RRGGBB) for openpyxl."""
    hex_clean = hex_color.replace('#', '').replace('0x', '')
    if len(hex_clean) == 6:
        return "00" + hex_clean.upper()
    return hex_clean.upper()


def generate_excel_report(start_date, end_date, occupied, available, occupancy_rate, type_rows, all_att, total_records, unique_p, s1_total, s2_total, act_summary, parts, parts_df, new_c, reg_c, unsigned):
    """Generate Excel report with multiple sheets matching the dashboard visuals."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()
    thin_border = Border(
        left=Side(style='thin', color='00CCCCCC'),
        right=Side(style='thin', color='00CCCCCC'),
        top=Side(style='thin', color='00CCCCCC'),
        bottom=Side(style='thin', color='00CCCCCC')
    )

    # ── Sheet 1: Summary Dashboard ──
    ws1 = wb.active
    ws1.title = "Summary"

    # Header
    ws1.merge_cells('A1:F1')
    ws1['A1'] = "Woodlands Zone 6 Community Hub - Monthly Meeting Report"
    ws1['A1'].font = Font(size=18, bold=True, color="00FFFFFF")
    ws1['A1'].fill = PatternFill(start_color="00667EEA", end_color="00667EEA", fill_type="solid")
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 30

    ws1.merge_cells('A2:F2')
    ws1['A2'] = f"Period: {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}"
    ws1['A2'].font = Font(size=12, italic=True)
    ws1['A2'].alignment = Alignment(horizontal='center')

    ws1.merge_cells('A3:F3')
    ws1['A3'] = f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}"
    ws1['A3'].font = Font(size=10, color="00666666")
    ws1['A3'].alignment = Alignment(horizontal='center')

    # Garden Section
    row = 5
    ws1.merge_cells(f'A{row}:F{row}')
    ws1[f'A{row}'] = "ROOF TOP GARDEN"
    ws1[f'A{row}'].font = Font(size=14, bold=True, color="00FFFFFF")
    ws1[f'A{row}'].fill = PatternFill(start_color="002CA02C", end_color="002CA02C", fill_type="solid")
    ws1[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws1.row_dimensions[row].height = 25

    row += 1
    garden_metrics = [
        ["Total Plots", TOTAL_PLOTS],
        ["Occupied", occupied],
        ["Available", available],
        ["Occupancy Rate", f"{occupancy_rate:.1%}"],
    ]
    for label, value in garden_metrics:
        ws1[f'A{row}'] = label
        ws1[f'B{row}'] = value
        ws1[f'A{row}'].font = Font(bold=True)
        ws1[f'B{row}'].font = Font(size=12, color="00667EEA", bold=True)
        row += 1

    # Plot Types Breakdown
    row += 1
    ws1.merge_cells(f'A{row}:F{row}')
    ws1[f'A{row}'] = "Plot Types Breakdown"
    ws1[f'A{row}'].font = Font(size=12, bold=True)
    row += 1

    headers = ["Plot Type", "Occupied", "Available", "Total", "Rate %"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True, color="00FFFFFF")
        cell.fill = PatternFill(start_color="00333333", end_color="00333333", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    row += 1

    for tr in type_rows:
        parts_type = tr["Plot Type"].split(" ")
        ptype = parts_type[1] if len(parts_type) > 1 else "B"
        color_hex = PLOT_TYPES.get(ptype, {}).get('colour', '#667EEA')
        argb_color = _hex_to_arbg(color_hex)

        ws1.cell(row=row, column=1, value=tr["Plot Type"])
        ws1.cell(row=row, column=2, value=tr["Occupied"])
        ws1.cell(row=row, column=3, value=tr["Available"])
        ws1.cell(row=row, column=4, value=tr["Occupied"] + tr["Available"])
        ws1.cell(row=row, column=5, value=f"{tr['Rate']:.1f}%")
        # Color the type cell background with plot type color
        ws1.cell(row=row, column=1).fill = PatternFill(start_color=argb_color, end_color=argb_color, fill_type="solid")
        ws1.cell(row=row, column=1).font = Font(color="00FFFFFF", bold=True)
        for c in range(1, 6):
            ws1.cell(row=row, column=c).border = thin_border
            ws1.cell(row=row, column=c).alignment = Alignment(horizontal='center')
        row += 1

    # Activity Section
    row += 1
    ws1.merge_cells(f'A{row}:F{row}')
    ws1[f'A{row}'] = "ACTIVITY PARTICIPATION"
    ws1[f'A{row}'].font = Font(size=14, bold=True, color="00FFFFFF")
    ws1[f'A{row}'].fill = PatternFill(start_color="00764BA2", end_color="00764BA2", fill_type="solid")
    ws1[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws1.row_dimensions[row].height = 25

    if all_att:
        row += 1
        activity_metrics = [
            ["Total Records", total_records],
            ["Unique Residents", int(unique_p)],
            ["Session 1", s1_total],
            ["Session 2", s2_total],
        ]
        for label, value in activity_metrics:
            ws1[f'A{row}'] = label
            ws1[f'B{row}'] = value
            ws1[f'A{row}'].font = Font(bold=True)
            ws1[f'B{row}'].font = Font(size=12, color="00667EEA", bold=True)
            row += 1

        if act_summary:
            row += 1
            ws1.merge_cells(f'A{row}:F{row}')
            ws1[f'A{row}'] = "By Activity"
            ws1[f'A{row}'].font = Font(size=12, bold=True)
            row += 1
            act_headers = ["Activity", "Records", "Unique Residents", "Session 1", "Session 2", "Both"]
            for col, h in enumerate(act_headers, 1):
                cell = ws1.cell(row=row, column=col, value=h)
                cell.font = Font(bold=True, color="00FFFFFF")
                cell.fill = PatternFill(start_color="00333333", end_color="00333333", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            row += 1
            for a in act_summary:
                ws1.cell(row=row, column=1, value=a['Activity'])
                ws1.cell(row=row, column=2, value=a['Records'])
                ws1.cell(row=row, column=3, value=a['Unique Residents'])
                ws1.cell(row=row, column=4, value=a['Session 1'])
                ws1.cell(row=row, column=5, value=a['Session 2'])
                ws1.cell(row=row, column=6, value=a['Both'])
                for c in range(1, 7):
                    ws1.cell(row=row, column=c).border = thin_border
                    ws1.cell(row=row, column=c).alignment = Alignment(horizontal='center')
                row += 1
    else:
        row += 1
        ws1[f'A{row}'] = "No attendance records for this period."
        ws1[f'A{row}'].font = Font(italic=True, color="00666666")

    # Resident Snapshot
    row += 1
    ws1.merge_cells(f'A{row}:F{row}')
    ws1[f'A{row}'] = "RESIDENT SNAPSHOT"
    ws1[f'A{row}'].font = Font(size=14, bold=True, color="00FFFFFF")
    ws1[f'A{row}'].fill = PatternFill(start_color="00FF7F0E", end_color="00FF7F0E", fill_type="solid")
    ws1[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws1.row_dimensions[row].height = 25

    try:
        if parts is not None and parts_df is not None:
            row += 1
            resident_metrics = [
                ["Total Registered", len(parts_df)],
                ["New", new_c],
                ["Regular", reg_c],
                ["Unsigned Indemnity", unsigned],
            ]
            for label, value in resident_metrics:
                ws1[f'A{row}'] = label
                ws1[f'B{row}'] = value
                ws1[f'A{row}'].font = Font(bold=True)
                ws1[f'B{row}'].font = Font(size=12, color="00667EEA", bold=True)
                row += 1
    except:
        row += 1
        ws1[f'A{row}'] = "Resident data unavailable."

    # Auto-fit columns
    for col in ws1.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws1.column_dimensions[column].width = adjusted_width

    # ── Sheet 2: Garden Plot Entitlements (Visual + Data) ──
    ws2 = wb.create_sheet("Garden Entitlements")

    ws2.merge_cells('A1:H1')
    ws2['A1'] = "Garden Plot Entitlements"
    ws2['A1'].font = Font(size=16, bold=True, color="00FFFFFF")
    ws2['A1'].fill = PatternFill(start_color="002CA02C", end_color="002CA02C", fill_type="solid")
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 25

    headers2 = ["Plot Number", "Plot Type", "Area (sqm)", "Status", "Owner ID", "Owner Name", "Contact", "Paid"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color="00FFFFFF")
        cell.fill = PatternFill(start_color="00333333", end_color="00333333", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Load participants for name lookup
    try:
        all_participants = supabase.table('participants').select("*").execute().data
        part_dict = {p['id'].lower().strip(): p for p in all_participants}
    except:
        part_dict = {}

    all_plots = load_plots()
    plots_dict = {p['plot_number']: p for p in all_plots}

    row = 3
    for pn in range(1, 77):
        ptype = TYPE_MAP.get(pn, 'B')
        area = PLOT_TYPES[ptype]["area"]
        plot = plots_dict.get(pn)

        if plot and plot.get('occupied'):
            status_text = "Occupied"
            owner_id = plot.get('user_id', '')
            resident = part_dict.get(str(owner_id).lower().strip())
            owner_name = plot.get('user_name', resident['name'] if resident else '')
            contact = plot.get('contact', resident.get('contact', '') if resident else '')
            paid_text = "Yes" if plot.get('paid') else "No"
            # Visual indicator: light red fill for occupied
            status_fill = PatternFill(start_color="00FFCCCC", end_color="00FFCCCC", fill_type="solid")
            status_font = Font(color="00CC0000", bold=True)
        else:
            status_text = "Available"
            owner_id = ""
            owner_name = ""
            contact = ""
            paid_text = ""
            # Visual indicator: light green fill for available
            status_fill = PatternFill(start_color="00CCFFCC", end_color="00CCFFCC", fill_type="solid")
            status_font = Font(color="00006600", bold=True)

        ws2.cell(row=row, column=1, value=pn)
        ws2.cell(row=row, column=2, value=ptype)
        ws2.cell(row=row, column=3, value=area)
        ws2.cell(row=row, column=4, value=status_text)
        ws2.cell(row=row, column=4).fill = status_fill
        ws2.cell(row=row, column=4).font = status_font
        ws2.cell(row=row, column=5, value=owner_id)
        ws2.cell(row=row, column=6, value=owner_name)
        ws2.cell(row=row, column=7, value=contact)
        ws2.cell(row=row, column=8, value=paid_text)
        for c in range(1, 9):
            ws2.cell(row=row, column=c).border = thin_border
            ws2.cell(row=row, column=c).alignment = Alignment(horizontal='center')
        row += 1

    # Auto-fit columns for sheet 2
    for col in ws2.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws2.column_dimensions[column].width = adjusted_width

    # ── Sheet 3: Raw Attendance Data ──
    if all_att:
        ws3 = wb.create_sheet("Attendance Data")
        att_df_export = pd.DataFrame(all_att)
        for r_idx, row_data in enumerate(dataframe_to_rows(att_df_export, index=False, header=True), 1):
            for c_idx, value in enumerate(row_data, 1):
                ws3.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    ws3.cell(row=r_idx, column=c_idx).font = Font(bold=True, color="00FFFFFF")
                    ws3.cell(row=r_idx, column=c_idx).fill = PatternFill(start_color="00333333", end_color="00333333", fill_type="solid")
                    ws3.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center')
                else:
                    ws3.cell(row=r_idx, column=c_idx).border = thin_border

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
